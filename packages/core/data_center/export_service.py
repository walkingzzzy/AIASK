"""
数据导出服务
"""
import uuid
import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import logging

from .models import DataQuery, DataExport, ExportFormat

logger = logging.getLogger(__name__)


class ExportService:
    """数据导出服务"""

    def __init__(self, export_dir: str = "data/exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self.exports: Dict[str, DataExport] = {}

    def create_export(self, query: DataQuery, format: ExportFormat) -> DataExport:
        """
        创建导出任务

        Args:
            query: 数据查询
            format: 导出格式

        Returns:
            导出任务
        """
        export_id = str(uuid.uuid4())
        export = DataExport(
            export_id=export_id,
            query=query,
            format=format,
            status="pending"
        )
        self.exports[export_id] = export
        return export

    def execute_export(self, export_id: str, data: List[Dict[str, Any]]) -> bool:
        """
        执行导出

        Args:
            export_id: 导出任务ID
            data: 要导出的数据

        Returns:
            是否成功
        """
        if export_id not in self.exports:
            return False

        export = self.exports[export_id]
        export.status = "processing"

        try:
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{export.query.category.value}_{timestamp}.{export.format.value}"
            file_path = self.export_dir / filename

            # 根据格式导出
            if export.format == ExportFormat.CSV:
                self._export_csv(file_path, data)
            elif export.format == ExportFormat.JSON:
                self._export_json(file_path, data)
            elif export.format == ExportFormat.EXCEL:
                self._export_excel(file_path, data)

            export.file_path = str(file_path)
            export.status = "completed"
            export.completed_at = datetime.now()
            return True

        except Exception as e:
            export.status = "failed"
            export.error_message = str(e)
            logger.error(f"导出失败: {e}")
            return False

    def _export_csv(self, file_path: Path, data: List[Dict[str, Any]]) -> None:
        """导出为CSV"""
        if not data:
            return

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

    def _export_json(self, file_path: Path, data: List[Dict[str, Any]]) -> None:
        """导出为JSON"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _export_excel(self, file_path: Path, data: List[Dict[str, Any]]) -> None:
        """导出为Excel"""
        if not data:
            return
        
        try:
            # 尝试使用 openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "数据导出"
            
            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
            
            # 自动调整列宽
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            cell_length = len(str(cell.value)) if cell.value else 0
                            max_length = max(max_length, cell_length)
                        except:
                            pass
                # 设置列宽（中文字符需要更宽）
                adjusted_width = min(max_length * 1.5 + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            wb.save(file_path)
            logger.info(f"Excel导出成功: {file_path}")
            
        except ImportError:
            # openpyxl 未安装，尝试使用 pandas
            try:
                import pandas as pd
                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False, engine='openpyxl')
                logger.info(f"使用pandas导出Excel: {file_path}")
            except ImportError:
                # 都不可用，降级为CSV
                logger.warning("openpyxl和pandas都未安装，降级为CSV格式")
                csv_path = file_path.with_suffix('.csv')
                self._export_csv(csv_path, data)
                # 更新文件路径
                return csv_path

    def get_export_status(self, export_id: str) -> Optional[DataExport]:
        """获取导出任务状态"""
        return self.exports.get(export_id)

    def list_exports(self) -> List[DataExport]:
        """列出所有导出任务"""
        return list(self.exports.values())

    def delete_export(self, export_id: str) -> bool:
        """删除导出任务"""
        if export_id not in self.exports:
            return False

        export = self.exports[export_id]
        if export.file_path and Path(export.file_path).exists():
            Path(export.file_path).unlink()

        del self.exports[export_id]
        return True
